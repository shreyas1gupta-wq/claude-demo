#!/usr/bin/env python3
"""Export the Demeter overview data to an Excel workbook (Demeter_Dual_Engine_Overview.xlsx).
Sheets: README, Monthly Returns, Monthly Matrix, Calendar Years, Statistics by Window,
Drawdowns, Worst & Best Months, Quadrants, Growth (data + native line chart)."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
A = json.loads((HERE / "analytics.json").read_text())
S = json.loads((HERE / "data" / "stated_figures.json").read_text())
FONT = "Arial"
NAVY = "14213D"
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=NAVY)
base = Font(name=FONT, size=10)
bold = Font(name=FONT, size=10, bold=True)
spy_font = Font(name=FONT, size=10, italic=True, color="9C6F12")
note_font = Font(name=FONT, size=9, italic=True, color="666666")
thin = Side(style="thin", color="DDDDDD")
PCT = '0.00%;-0.00%;"-"'
PCT1 = '0.0%'
NUM3 = '0.000'
MONEY = '"$"#,##0'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths=None):
    for i, col in enumerate(ws.columns, start=1):
        w = (widths or {}).get(i)
        if w is None:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            w = min(max(10, w + 2), 60)
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_readme(wb):
    ws = wb.active; ws.title = "README"
    lines = [
        ("Demeter Tactical Investments - Dual-Engine Quantitative Equity Strategy (Share Class B)", bold),
        (f"Data: {A['meta']['n_months']} published monthly returns, {A['meta']['inception']} to {A['meta']['as_of']}, for the strategy and SPY (SPDR S&P 500 ETF Trust, total return).", base),
        ("Source: Demeter factsheet tables, transcribed by Ionic Wealth research. Compounding the transcribed months reproduces every published YTD within 0.02 pp.", base),
        ("Sheets:", bold),
        ("  Monthly Returns  - long format, one row per month (percent as fractions, e.g. 0.0467 = 4.67%).", base),
        ("  Monthly Matrix   - year x month grid with the compounded year alongside the published figure.", base),
        ("  Calendar Years   - strategy vs SPY per year with excess return.", base),
        ("  Statistics       - the factsheet's four windows plus the two live periods, recomputed with the factsheet's conventions.", base),
        ("  Drawdowns        - five largest peak-to-trough episodes per series.", base),
        ("  Best Worst       - ten best/worst months per series with the other series alongside.", base),
        ("  Quadrants        - Demeter's four-quadrant daily framework as published (note the 3,219 vs 3,477 day discrepancy).", base),
        ("  Growth           - growth of $1,000 and drawdowns by month, with a line chart.", base),
        ("Conventions (from analytics.json meta.definitions):", bold),
    ]
    for k, v in A["meta"]["definitions"].items():
        lines.append((f"  {k}: {v}", note_font))
    lines.append((f"Risk-free rate: {A['meta']['tb3ms_source']} {A['meta']['tb3ms_note']}", note_font))
    lines.append(("All cells are values computed by build_analytics.py / build_workbook.py (no live formulas), so the workbook opens with every number populated in any viewer.", note_font))
    lines.append(("Past performance is not indicative of future results. Leverage magnifies losses as well as gains.", note_font))
    for i, (t, f) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=t); c.font = f; c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 140


def sheet_monthly(wb):
    ws = wb.create_sheet("Monthly Returns")
    ws.append(["Month", "Period", "Strategy", "SPY", "Excess vs SPY", "TB3MS (ann. %)", "Growth $1,000 Strategy", "Growth $1,000 SPY", "Drawdown Strategy", "Drawdown SPY"])
    s = A["series"]
    for i in range(len(s["months"])):
        r = i + 2
        ws.append([s["months"][i], s["period"][i], s["strategy"][i] / 100, s["spy"][i] / 100, (s["strategy"][i] - s["spy"][i]) / 100, s["tb3ms_pct"][i] / 100,
                   s["growth_strategy"][i], s["growth_spy"][i], s["dd_strategy"][i] / 100, s["dd_spy"][i] / 100])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = base
        for c in (row[2], row[3], row[4], row[8], row[9]): c.number_format = PCT
        row[5].number_format = '0.00%'
        row[6].number_format = MONEY; row[7].number_format = MONEY
    style_header(ws, 1, 10); ws.freeze_panes = "A2"; autosize(ws, {1: 10, 2: 16})
    return ws


def sheet_matrix(wb):
    ws = wb.create_sheet("Monthly Matrix")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    ws.append(["Year", "Series"] + months + ["Year (compounded)", "Year published"])
    stated = S["period_ytd_stated"]
    r = 2
    for row in A["monthly_table"]:
        for series, key, font in (("Strategy", "strategy", bold), ("SPY", "spy", spy_font)):
            vals = [None if v is None else v / 100 for v in row[key]]
            first = get_column_letter(3); last = get_column_letter(14)
            ytd = 1.0
            for v in vals:
                if v is not None: ytd *= 1 + v
            ws.append([row["year"] if series == "Strategy" else None, series] + vals + [ytd - 1, None])
            pub = stated[key].get(str(row["year"]))
            if pub is None and row["year"] == 2023:
                pub = None  # published as two periods; see Calendar Years
            if pub is None and row["year"] == 2026:
                pub = stated[key].get("2026_jan_jun")
            ws.cell(row=r, column=16, value=None if pub is None else pub / 100)
            for c in range(1, 17):
                cell = ws.cell(row=r, column=c); cell.font = font
                if c >= 3: cell.number_format = PCT
            r += 1
    ws.cell(row=r + 1, column=1, value="Year (compounded) = product of (1 + monthly return) - 1 over the available months; 2012 is Jul-Dec and 2026 is Jan-Jun. 2023 was published as two periods (Jan-Jul 26.43% / 20.58%; Aug-Dec -0.87% / 4.63%), so its full-year figure has no published counterpart.").font = note_font
    style_header(ws, 1, 16); ws.freeze_panes = "C2"; autosize(ws, {1: 8, 2: 10, 15: 14, 16: 14})


def sheet_years(wb):
    ws = wb.create_sheet("Calendar Years")
    ws.append(["Year", "Months", "Partial", "Strategy", "SPY", "Excess", "Strategy best month", "Strategy worst month", "Strategy positive months", "SPY positive months", "Months strategy beat SPY"])
    for i, y in enumerate(A["calendar_years"], start=2):
        ws.append([y["year"], y["n_months"], "yes" if y["partial"] else "", y["strategy_pct"] / 100, y["spy_pct"] / 100, y["excess_pct"] / 100, y["strategy_best_month_pct"] / 100, y["strategy_worst_month_pct"] / 100, y["strategy_pos_months"], y["spy_pos_months"], y["months_strategy_beat_spy"]])
    n = len(A["calendar_years"]) + 1
    for row in ws.iter_rows(min_row=2, max_row=n):
        for c in row: c.font = base
        for c in (row[3], row[4], row[5], row[6], row[7]): c.number_format = PCT
    ws.cell(row=n + 2, column=1, value="Full years beating SPY:").font = bold
    ws.cell(row=n + 2, column=4, value=sum(1 for y in A["calendar_years"] if not y["partial"] and y["excess_pct"] > 0)).font = base
    ws.cell(row=n + 3, column=1, value="Full years in table:").font = bold
    ws.cell(row=n + 3, column=4, value=sum(1 for y in A["calendar_years"] if not y["partial"])).font = base
    style_header(ws, 1, 11); ws.freeze_panes = "A2"; autosize(ws)


def sheet_stats(wb):
    ws = wb.create_sheet("Statistics")
    metrics = [("total_growth_pct", "Total growth", "pct"), ("annualized_return_pct", "Annualised return", "pct"), ("annualized_std_dev", "Annualised std. dev.", "num"), ("max_drawdown_pct", "Maximum drawdown", "pct"), ("pct_positive_months", "% positive months", "pct1"), ("sharpe", "Sharpe ratio", "num"), ("sortino", "Sortino ratio", "num"), ("calmar", "Calmar ratio", "num"), ("correlation_to_spy", "Correlation to SPY", "num"), ("beta_to_spy", "Beta to SPY", "num"), ("alpha_ann_jensen_pct", "Alpha (ann., Jensen)", "pct"), ("up_capture_pct", "Up capture vs SPY", "pct1"), ("down_capture_pct", "Down capture vs SPY", "pct1"), ("growth_of_1000", "Growth of $1,000", "money"), ("best_month_pct", "Best month", "pct"), ("worst_month_pct", "Worst month", "pct"), ("n_up_months", "Up months", "int"), ("n_down_months", "Down months", "int")]
    wins = [("since_inception", "Since inception"), ("rolling_120m", "Rolling 120 months"), ("rolling_60m", "Rolling 60 months"), ("rolling_12m", "Rolling 12 months"), ("family_office", "Family-office period"), ("commingled_fund", "Commingled fund")]
    hdr = ["Statistic"]
    for k, l in wins:
        w = A["windows"][k]; hdr += [f"{l}\n{w['start']} to {w['end']}\nStrategy", "SPY", "S&P 500 price (published)"]
    ws.append(hdr)
    for mk, ml, kind in metrics:
        row = [ml]
        for k, _ in wins:
            w = A["windows"][k]
            vs, vp = w["strategy"].get(mk), w["spy"].get(mk)
            st = (w.get("sp500_price_stated") or {}).get(mk)
            conv = (lambda v: None if v is None else v / 100) if kind in ("pct", "pct1") else (lambda v: v)
            row += [conv(vs), conv(vp), conv(st)]
        ws.append(row)
        r = ws.max_row
        for c in range(2, 2 + 3 * len(wins)):
            cell = ws.cell(row=r, column=c); cell.font = base
            cell.number_format = {"pct": PCT, "pct1": PCT1, "num": NUM3, "money": MONEY, "int": "0"}[kind]
            if (c - 2) % 3 == 0: cell.font = bold
            if (c - 2) % 3 == 2: cell.font = Font(name=FONT, size=10, color="888888")
        ws.cell(row=r, column=1).font = base
    rs = A.get("reconciliation_summary", {})
    ws.cell(row=ws.max_row + 2, column=1, value=f"Reconciliation to the published factsheet: {rs.get('n_ok')} of {rs.get('n_checks')} statistics within tolerance; the {rs.get('n_fail')} exceptions are SPY total-growth figures off by 0.06-0.23 pp from compounding 2-decimal monthly returns. S&P 500 price column is quoted from the factsheet (price-only index; no monthly series supplied).").font = note_font
    style_header(ws, 1, 1 + 3 * len(wins)); ws.row_dimensions[1].height = 48; ws.freeze_panes = "B2"; autosize(ws, {1: 24})


def sheet_drawdowns(wb):
    ws = wb.create_sheet("Drawdowns")
    ws.append(["Series", "Depth", "Peak", "Trough", "Months to trough", "Recovered", "Months to recover"])
    for series in ("strategy", "spy"):
        for d in A["drawdowns"][series]:
            ws.append([series.upper() if series == "spy" else "Strategy", d["depth_pct"] / 100, d["peak"], d["trough"], d["length_months"], d["recovery"] or "not yet", d["recovery_months"]])
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = spy_font if row[0].value == "SPY" else base
        row[1].number_format = PCT
    style_header(ws, 1, 7); autosize(ws)


def sheet_bestworst(wb):
    ws = wb.create_sheet("Best Worst")
    bw = A["best_worst"]
    blocks = [("Strategy's ten worst months", bw["strategy_worst_10"]), ("Strategy's ten best months", bw["strategy_best_10"]), ("SPY's ten worst months", bw["spy_worst_10"]), ("SPY's ten best months", bw["spy_best_10"])]
    r = 1
    for title, rows in blocks:
        ws.cell(row=r, column=1, value=title).font = bold; r += 1
        for c, hname in enumerate(["Month", "Strategy", "SPY", "Excess"], start=1):
            ws.cell(row=r, column=c, value=hname)
        style_header(ws, r, 4); r += 1
        for x in rows:
            ws.cell(row=r, column=1, value=x["month"]).font = base
            ws.cell(row=r, column=2, value=x["strategy"] / 100).font = base
            ws.cell(row=r, column=3, value=x["spy"] / 100).font = base
            ws.cell(row=r, column=4, value=(x["strategy"] - x["spy"]) / 100).font = base
            for c in (2, 3, 4): ws.cell(row=r, column=c).number_format = PCT
            r += 1
        r += 1
    ud = A["up_down_analysis"]
    ws.cell(row=r, column=1, value="SPY-down months: strategy behaviour").font = bold; r += 1
    for lab, val, fmt in [("SPY down months", ud["n_spy_down_months"], "0"), ("Strategy average in SPY-down months", ud["strategy_avg_in_spy_down_pct"] / 100, PCT), ("SPY average in its down months", ud["spy_avg_in_down_pct"] / 100, PCT), ("Strategy beat SPY in down months", ud["strategy_beat_spy_in_down_months_pct"] / 100, PCT1), ("Strategy positive in SPY-down months", ud["strategy_positive_when_spy_down_pct"] / 100, PCT1), ("Months strategy beat SPY (all)", ud["months_strategy_beat_spy"], "0")]:
        ws.cell(row=r, column=1, value=lab).font = base; c = ws.cell(row=r, column=2, value=val); c.font = base; c.number_format = fmt; r += 1
    autosize(ws, {1: 40})


def sheet_quadrants(wb):
    ws = wb.create_sheet("Quadrants")
    q = S["quadrants_daily"]
    ws.append(["Quadrant", "Days (published)", "Share of quadrant sum", "Description"])
    total = q["loss_avoidance_days"] + q["gain_sacrifice_days"] + q["amplified_gains_days"] + q["amplified_losses_days"]
    rows = [("Loss avoidance", q["loss_avoidance_days"], "In cash while the market fell"), ("Gain sacrifice", q["gain_sacrifice_days"], "In cash while the market rose"), ("Amplified gains", q["amplified_gains_days"], "Invested (levered) while the market rose"), ("Amplified losses", q["amplified_losses_days"], "Invested (levered) while the market fell")]
    for i, (l, d, desc) in enumerate(rows, start=2):
        ws.append([l, d, d / total, desc])
        for c in ws[i]: c.font = base
        ws.cell(row=i, column=3).number_format = PCT1
    ws.append(["Sum of quadrants", total, None, "Published total trading days: %d" % q["stated_total_days"]])
    ws.append(["Cash days", q["loss_avoidance_days"] + q["gain_sacrifice_days"], (q["loss_avoidance_days"] + q["gain_sacrifice_days"]) / total, "Manager quotes ~50% of trading days in cash"])
    ws.append(["Positive days / published total", q["amplified_gains_days"] / q["stated_total_days"], None, "Manager quotes ~31% positive days (1,022 / 3,219)"])
    for r in (6, 7, 8):
        for c in ws[r]: c.font = base
    ws["B8"].number_format = PCT1; ws["C7"].number_format = PCT1
    ws.cell(row=10, column=1, value=q["_source_note"]).font = note_font
    style_header(ws, 1, 4); autosize(ws, {1: 30, 4: 60})


def sheet_growth(wb):
    ws = wb.create_sheet("Growth")
    s = A["series"]
    ws.append(["Month", "Growth $1,000 Strategy", "Growth $1,000 SPY", "Drawdown Strategy", "Drawdown SPY", "Rolling 12m Strategy", "Rolling 12m SPY"])
    for i in range(len(s["months"])):
        ws.append([s["months"][i], s["growth_strategy"][i], s["growth_spy"][i], s["dd_strategy"][i] / 100, s["dd_spy"][i] / 100,
                   None if s["rolling12_strategy"][i] is None else s["rolling12_strategy"][i] / 100, None if s["rolling12_spy"][i] is None else s["rolling12_spy"][i] / 100])
    n = ws.max_row
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = base
        row[1].number_format = MONEY; row[2].number_format = MONEY
        for c in row[3:]: c.number_format = PCT
    style_header(ws, 1, 7); ws.freeze_panes = "A2"; autosize(ws, {1: 10})
    ch = LineChart(); ch.title = "Growth of $1,000 (log scale)"; ch.height = 11; ch.width = 26
    ch.y_axis.scaling.logBase = 10; ch.y_axis.title = "$"; ch.x_axis.title = "Month"
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_col=3, max_row=n), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
    ch.x_axis.tickLblSkip = 12; ch.x_axis.number_format = "@"
    ch.series[0].graphicalProperties.line.solidFill = "2748C6"; ch.series[1].graphicalProperties.line.solidFill = "B7861B"
    for sr in ch.series: sr.smooth = False; sr.graphicalProperties.line.width = 20000
    ws.add_chart(ch, "I2")
    ch2 = LineChart(); ch2.title = "Drawdown from prior peak"; ch2.height = 9; ch2.width = 26
    ch2.add_data(Reference(ws, min_col=4, min_row=1, max_col=5, max_row=n), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n)); ch2.x_axis.tickLblSkip = 12; ch2.y_axis.number_format = "0%"
    ch2.series[0].graphicalProperties.line.solidFill = "2748C6"; ch2.series[1].graphicalProperties.line.solidFill = "B7861B"
    for sr in ch2.series: sr.smooth = False; sr.graphicalProperties.line.width = 20000
    ws.add_chart(ch2, "I26")


def main():
    wb = Workbook()
    sheet_readme(wb); sheet_monthly(wb); sheet_matrix(wb); sheet_years(wb); sheet_stats(wb); sheet_drawdowns(wb); sheet_bestworst(wb); sheet_quadrants(wb); sheet_growth(wb)
    out = HERE / "Demeter_Dual_Engine_Overview.xlsx"; wb.save(out); print("wrote", out)


if __name__ == "__main__":
    main()
