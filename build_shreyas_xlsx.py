#!/usr/bin/env python3
"""Build Ionic Wealth client portfolio Excel for Shreyas Gupta.
15-holding equal-weight book: 11 stocks + 4 mutual funds.

Scores/recommendations are left as 'Pending scoring run' — the ionic-scorecard
pipeline is not available in this session and the skill forbids fabricating them.

LibreOffice/recalc is not runnable in this sandbox, so we inject cached values for
every (deterministic) formula cell directly into the sheet XML: the workbook keeps
LIVE formulas and also displays correct numbers in any previewer.
"""

import re, zipfile, shutil, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- house palette (from template/azby_data.json) ----
NAVY="16233B"; INDIGO="1B27A3"; INDIGO_T="8C95DE"; SLATE="6B7280"
GRID="E5E7EB"; TINT="EEEFF7"; NEAR="F5F6FC"; WHITE="FFFFFF"
FONT="Arial"; PENDING="Pending scoring run"

STOCKS = [
    ("KPI Green Energy",         "KPIGREEN",  "Power / Renewables",           "Small-Mid"),
    ("Infosys",                  "INFY",      "Information Technology",       "Large"),
    ("Reliance Industries",      "RELIANCE",  "Oil, Gas & Consumable Fuels",  "Large"),
    ("KPIT Technologies",        "KPITTECH",  "Information Technology",       "Mid"),
    ("Jyoti Resins & Adhesives", "JYOTIRES",  "Chemicals",                    "Small"),
    ("Kalyan Jewellers India",   "KALYANKJIL","Consumer Durables",            "Large-Mid"),
    ("Glenmark Pharmaceuticals", "GLENMARK",  "Pharmaceuticals & Biotech",    "Mid"),
    ("Vodafone Idea",            "IDEA",      "Telecommunication",            "Mid"),
    ("Anant Raj",                "ANANTRAJ",  "Realty",                       "Mid"),
    ("Godrej Industries",        "GODREJIND", "Diversified / Chemicals",      "Large-Mid"),
    ("Cupid",                    "CUPID",     "Healthcare (Medical Devices)", "Small"),
]
FUNDS = [
    ("WhiteOak Capital Multi Asset Allocation Fund", "WhiteOak", "Direct",  "Multi Asset Allocation (Hybrid)"),
    ("LIC MF Small Cap Fund",                        "LIC MF",   "Regular", "Small Cap (Equity)"),
    ("Nippon India Small Cap Fund",                  "Nippon",   "Regular", "Small Cap (Equity)"),
    ("HSBC Large Cap Fund",                          "HSBC",     "Regular", "Large Cap (Equity)"),
]
N = len(STOCKS) + len(FUNDS)  # 15

# ---- styling helpers ----
def fill(h): return PatternFill("solid", fgColor=h)
thin = Side(style="thin", color=GRID)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

# cached-value recorder: {sheet_title: {coord: number}}
VALUES = {}

def style_header(cell):
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=10)
    cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = box

def cs(ws, coord, val, *, bold=False, color="000000", size=10, fillhex=None,
       align="left", num=None, border=True, italic=False, wrap=False, cached=None):
    c = ws[coord]; c.value = val
    c.font = Font(name=FONT, bold=bold, color=color, size=size, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fillhex: c.fill = fill(fillhex)
    if num: c.number_format = num
    if border: c.border = box
    if cached is not None:
        VALUES.setdefault(ws.title, {})[coord] = cached
    return c

wb = Workbook()

# ===================== SHEET 1 — At a Glance =====================
s1 = wb.active; s1.title = "At a Glance"; s1.sheet_view.showGridLines = False
for col,w in {"A":2,"B":34,"C":18,"D":18,"E":22,"F":2}.items():
    s1.column_dimensions[col].width = w

cs(s1,"B2","IONIC WEALTH  |  Portfolio Snapshot",bold=True,color=NAVY,size=16,border=False)
cs(s1,"B3","Client: Shreyas Gupta   ·   Account: NDPMS   ·   As of: 30 July 2026",color=SLATE,border=False)

for col,label,val,cached in [("B","Total Holdings","=15",15),("C","Equity (Stocks)","=11",11),("D","Mutual Funds","=4",4)]:
    cs(s1,f"{col}5",label,bold=True,color=WHITE,fillhex=INDIGO,align="center")
    cs(s1,f"{col}6",val,bold=True,color=NAVY,fillhex=TINT,size=18,align="center",num="0",cached=cached)

cs(s1,"B8","Construction",bold=True,color=WHITE,fillhex=NAVY)
s1.merge_cells("C8:D8"); cs(s1,"C8","Equal weight across all 15 holdings",fillhex=NEAR)
cs(s1,"E8","",fillhex=NEAR,border=True)
cs(s1,"B9","Target weight per holding",bold=True,color=WHITE,fillhex=NAVY)
s1.merge_cells("C9:D9"); cs(s1,"C9","=1/15",color=NAVY,bold=True,fillhex=NEAR,num="0.00%",align="center",cached=1/15)
cs(s1,"E9","",fillhex=NEAR,border=True)

cs(s1,"B11","Allocation by sleeve",bold=True,color=NAVY,size=12,border=False)
for j,h in enumerate(["Sleeve","# Holdings","Weight","Basis"]):
    style_header(s1[f"{get_column_letter(2+j)}12"]); s1[f"{get_column_letter(2+j)}12"]=h
cs(s1,"B13","Direct Equity",bold=True,fillhex=NEAR)
cs(s1,"C13","=11",align="center",num="0",cached=11)
cs(s1,"D13","=C13/15",align="center",num="0.00%",cached=11/15)
cs(s1,"E13","11 x equal weight",color=SLATE)
cs(s1,"B14","Mutual Funds",bold=True,fillhex=NEAR)
cs(s1,"C14","=4",align="center",num="0",cached=4)
cs(s1,"D14","=C14/15",align="center",num="0.00%",cached=4/15)
cs(s1,"E14","4 x equal weight",color=SLATE)
cs(s1,"B15","Total",bold=True,color=NAVY,fillhex=INDIGO_T)
cs(s1,"C15","=SUM(C13:C14)",bold=True,align="center",num="0",color=NAVY,fillhex=INDIGO_T,cached=15)
cs(s1,"D15","=SUM(D13:D14)",bold=True,align="center",num="0.00%",color=NAVY,fillhex=INDIGO_T,cached=1.0)
cs(s1,"E15","",fillhex=INDIGO_T)

cs(s1,"B17","Data status",bold=True,color=NAVY,size=12,border=False)
s1.merge_cells("B18:E18")
cs(s1,"B18",
   "Ionic Scores and Sell / Trim / Hold recommendations are marked 'Pending scoring run'. "
   "The scorecard pipeline is not available in this session; no scores or verdicts have been "
   "fabricated. ISINs and rupee values are left blank for the RM to populate from the CAS statement.",
   fillhex="FFF7E6",wrap=True)
for c in ("C18","D18","E18"): s1[c].fill=fill("FFF7E6"); s1[c].border=box
s1.row_dimensions[18].height = 60

# ===================== SHEET 2 — Recommendations =====================
s2 = wb.create_sheet("Recommendations"); s2.sheet_view.showGridLines=False
headers2=["#","Holding","Ticker","ISIN","Sector","Mkt-Cap","Weight","Ionic Score","Recommendation","Trim-to","Rationale"]
widths2=[4,30,12,16,30,11,10,12,16,10,34]
for j,(h,w) in enumerate(zip(headers2,widths2)):
    col=get_column_letter(1+j); s2.column_dimensions[col].width=w
    style_header(s2[f"{col}1"]); s2[f"{col}1"]=h

r=2
for i,(name,tick,sector,mcap) in enumerate(STOCKS,start=1):
    cs(s2,f"A{r}",i,align="center")
    cs(s2,f"B{r}",name,bold=True)
    cs(s2,f"C{r}",tick,align="center",color=SLATE)
    cs(s2,f"D{r}","—",align="center",color=SLATE)
    cs(s2,f"E{r}",sector)
    cs(s2,f"F{r}",mcap,align="center")
    cs(s2,f"G{r}","=1/15",align="center",num="0.00%",cached=1/15)
    cs(s2,f"H{r}",PENDING,italic=True,color=SLATE,align="center",size=9)
    cs(s2,f"I{r}",PENDING,italic=True,color=SLATE,align="center",size=9)
    cs(s2,f"J{r}","—",align="center",color=SLATE)
    cs(s2,f"K{r}",PENDING,italic=True,color=SLATE,size=9)
    r+=1
for k,(scheme,amc,plan,cat) in enumerate(FUNDS,start=1):
    cs(s2,f"A{r}",len(STOCKS)+k,align="center",fillhex=NEAR)
    cs(s2,f"B{r}",f"{scheme} — {plan}",bold=True,fillhex=NEAR)
    cs(s2,f"C{r}",amc,align="center",color=SLATE,fillhex=NEAR)
    cs(s2,f"D{r}","—",align="center",color=SLATE,fillhex=NEAR)
    cs(s2,f"E{r}",cat,fillhex=NEAR)
    cs(s2,f"F{r}","Fund",align="center",fillhex=NEAR)
    cs(s2,f"G{r}","=1/15",align="center",num="0.00%",fillhex=NEAR,cached=1/15)
    cs(s2,f"H{r}",PENDING,italic=True,color=SLATE,align="center",size=9,fillhex=NEAR)
    cs(s2,f"I{r}",PENDING,italic=True,color=SLATE,align="center",size=9,fillhex=NEAR)
    cs(s2,f"J{r}","—",align="center",color=SLATE,fillhex=NEAR)
    cs(s2,f"K{r}",PENDING,italic=True,color=SLATE,size=9,fillhex=NEAR)
    r+=1
last=r-1
cs(s2,f"A{r}","",fillhex=INDIGO_T)
cs(s2,f"B{r}","Total",bold=True,color=NAVY,fillhex=INDIGO_T)
for col in ["C","D","E","F"]: cs(s2,f"{col}{r}","",fillhex=INDIGO_T)
cs(s2,f"G{r}",f"=SUM(G2:G{last})",bold=True,color=NAVY,fillhex=INDIGO_T,align="center",num="0.00%",cached=1.0)
for col in ["H","I","J","K"]: cs(s2,f"{col}{r}","",fillhex=INDIGO_T)

r+=2
cs(s2,f"B{r}","Legend",bold=True,color=NAVY,size=11,border=False)
for note in [
    "Weight — equal weight = 1/15 of the book per holding (live formula; total = 100%).",
    "ISIN — left blank ( — ); to be filled by the RM from the NSDL CAS statement.",
    "Mkt-Cap tier — [INFERENCE] from general knowledge; verify against a live data feed.",
    "Ionic Score / Recommendation / Trim-to — pending a scorecard run; never fabricated.",
    "When scored, vocabulary is Sell / Trim / Hold only (never Buy) — this reviews existing holdings.",
]:
    r+=1; cs(s2,f"B{r}","•  "+note,color=SLATE,size=9,border=False)
s2.freeze_panes="A2"

# ===================== SHEET 3 — Composition =====================
s3=wb.create_sheet("Composition"); s3.sheet_view.showGridLines=False
for col,w in {"A":2,"B":38,"C":12,"D":12,"E":2}.items(): s3.column_dimensions[col].width=w
cs(s3,"B2","Portfolio Composition (equal-weight basis)",bold=True,color=NAVY,size=14,border=False)
cs(s3,"B3","Each holding = 1/15 = 6.67%. Sector is reference data; market-cap tier is [INFERENCE].",color=SLATE,size=9,border=False)

from collections import Counter
sec_counts=Counter(s[2] for s in STOCKS)
cs(s3,"B5","Equity by sector",bold=True,color=NAVY,size=12,border=False)
for j,h in enumerate(["Sector","# Stocks","Weight"]):
    col=get_column_letter(2+j); style_header(s3[f"{col}6"]); s3[f"{col}6"]=h
r=7; sec_start=r
for sec,cnt in sorted(sec_counts.items(),key=lambda x:(-x[1],x[0])):
    cs(s3,f"B{r}",sec)
    cs(s3,f"C{r}",cnt,align="center",num="0")
    cs(s3,f"D{r}",f"=C{r}/15",align="center",num="0.00%",cached=cnt/15)
    r+=1
cs(s3,f"B{r}","Mutual Funds (all schemes)",bold=True,fillhex=NEAR)
cs(s3,f"C{r}",4,align="center",num="0",fillhex=NEAR)
cs(s3,f"D{r}",f"=C{r}/15",align="center",num="0.00%",fillhex=NEAR,cached=4/15)
r+=1
cs(s3,f"B{r}","Total",bold=True,color=NAVY,fillhex=INDIGO_T)
cs(s3,f"C{r}",f"=SUM(C{sec_start}:C{r-1})",bold=True,align="center",num="0",color=NAVY,fillhex=INDIGO_T,cached=15)
cs(s3,f"D{r}",f"=SUM(D{sec_start}:D{r-1})",bold=True,align="center",num="0.00%",color=NAVY,fillhex=INDIGO_T,cached=1.0)

mcap_counts=Counter(s[3] for s in STOCKS)
r+=3
cs(s3,f"B{r}","Equity by market-cap tier  [INFERENCE — verify]",bold=True,color=NAVY,size=12,border=False)
r+=1
for j,h in enumerate(["Mkt-Cap tier","# Stocks","Weight"]):
    col=get_column_letter(2+j); style_header(s3[f"{col}{r}"]); s3[f"{col}{r}"]=h
r+=1; cap_start=r
for cap,cnt in sorted(mcap_counts.items(),key=lambda x:(-x[1],x[0])):
    cs(s3,f"B{r}",cap)
    cs(s3,f"C{r}",cnt,align="center",num="0")
    cs(s3,f"D{r}",f"=C{r}/15",align="center",num="0.00%",cached=cnt/15)
    r+=1
cs(s3,f"B{r}","Total (equity)",bold=True,color=NAVY,fillhex=INDIGO_T)
cs(s3,f"C{r}",f"=SUM(C{cap_start}:C{r-1})",bold=True,align="center",num="0",color=NAVY,fillhex=INDIGO_T,cached=11)
cs(s3,f"D{r}",f"=SUM(D{cap_start}:D{r-1})",bold=True,align="center",num="0.00%",color=NAVY,fillhex=INDIGO_T,cached=11/15)

OUT="/home/user/claude-demo/Ionic_Portfolio_Review_Shreyas_Gupta.xlsx"
wb.save(OUT)

# ---- inject cached values into formula cells (LibreOffice unavailable) ----
sheet_order=[ws.title for ws in wb.worksheets]           # sheet1.xml, sheet2.xml, ...
tmp=OUT+".tmp"
with zipfile.ZipFile(OUT) as zin, zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED) as zout:
    for item in zin.namelist():
        data=zin.read(item)
        m=re.match(r"xl/worksheets/sheet(\d+)\.xml$",item)
        if m:
            idx=int(m.group(1))-1
            title=sheet_order[idx] if idx < len(sheet_order) else None
            cmap=VALUES.get(title,{})
            if cmap:
                xml=data.decode("utf-8")
                for coord,num in cmap.items():
                    vnum=repr(float(num))
                    # <c r="COORD" ...><f>...</f></c>  ->  insert <v>NUM</v> before </c>
                    pat=re.compile(r'(<c r="'+re.escape(coord)+r'"[^>]*><f>[^<]*</f>)<v\s*/>(</c>)')
                    xml,nsub=pat.subn(r'\1<v>'+vnum+r'</v>\2',xml)
                    if nsub==0:
                        raise SystemExit(f"WARN: no formula cell matched for {title}!{coord}")
                data=xml.encode("utf-8")
        zout.writestr(item,data)
shutil.move(tmp,OUT)
print("WROTE",OUT,"| injected cached values for",sum(len(v) for v in VALUES.values()),"formula cells")
