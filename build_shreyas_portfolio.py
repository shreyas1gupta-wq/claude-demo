"""
Build the Shreyas Gupta portfolio workbook (Ionic Wealth client-facing format).

Honest-data build: holdings, sectors, and equal-weight structure are populated.
Ionic Scores / recommendations / fund QFRA verdicts are LEFT BLANK on purpose —
they require the Stock Scorecard 750 + QFRA pipeline (quant CSV + analyst JSON),
which is not run here. Per the manual's no-fabrication rule, we do not invent them.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- House style (from azby_data.json palette) --------------------------------
NAVY   = "16233B"
INDIGO = "1B27A3"
SLATE  = "6B7280"
GRID   = "E5E7EB"
NEAR   = "F5F6FC"
TINT_A = "EEEFF7"   # indigo tint (header bands)
AMBER  = "F2A93C"
YELLOWFILL = "FFF7E0"

FONT = "Arial"
CLIENT = "Shreyas Gupta"
AS_OF  = "30 July 2026"

# ---- Data ---------------------------------------------------------------------
# (stock, full company name, sector classification)
STOCKS = [
    ("KPI GREEN",    "KPI Green Energy Ltd",            "Power / Renewable Energy"),
    ("INFOSYS",      "Infosys Ltd",                     "Information Technology"),
    ("RELIANCE",     "Reliance Industries Ltd",         "Oil, Gas & Consumable Fuels"),
    ("KPIT TECH",    "KPIT Technologies Ltd",           "Information Technology"),
    ("JYOTI RESINS", "Jyoti Resins & Adhesives Ltd",    "Chemicals"),
    ("KALYAN JEWEL", "Kalyan Jewellers India Ltd",      "Consumer Durables"),
    ("GLENMARK",     "Glenmark Pharmaceuticals Ltd",    "Pharma & Healthcare"),
    ("VODAFONE IDEA","Vodafone Idea Ltd",               "Telecommunication"),
    ("ANANT RAJ",    "Anant Raj Ltd",                   "Realty"),
    ("GODREJ IND",   "Godrej Industries Ltd",           "Diversified / Chemicals"),
]

# (scheme, AMC, SEBI category, plan)
FUNDS = [
    ("WhiteOak Capital Multi Asset Allocation Fund", "WhiteOak Capital", "Multi Asset Allocation", "Direct"),
    ("LIC MF Small Cap Fund",                        "LIC Mutual Fund",  "Small Cap",             "Regular"),
    ("Nippon India Small Cap Fund",                  "Nippon India",     "Small Cap",             "Not stated"),
    ("HSBC Large Cap Fund",                          "HSBC Mutual Fund", "Large Cap",             "Not stated"),
]

N_HOLD = len(STOCKS) + len(FUNDS)   # 14 -> equal weight = 1/14 each

# ---- Style helpers ------------------------------------------------------------
thin = Side(style="thin", color=GRID)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, ref, val, *, bold=False, size=10, color="000000", fill=None,
         align="left", wrap=False, border=False, num=None, italic=False):
    c = ws[ref]
    c.value = val
    c.font = Font(name=FONT, bold=bold, italic=italic, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if border:
        c.border = box
    if num:
        c.number_format = num
    return c

def header_band(ws, row, cols, labels):
    for j, lab in enumerate(labels):
        ref = f"{get_column_letter(cols[0]+j)}{row}"
        cell(ws, ref, lab, bold=True, size=10, color="FFFFFF", fill=NAVY,
             align="center", wrap=True, border=True)

wb = openpyxl.Workbook()

# ============================================================ SHEET 1: SUMMARY
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
for col, w in {"A":2, "B":30, "C":26, "D":26, "E":22, "F":18}.items():
    ws.column_dimensions[col].width = w

cell(ws, "B2", "IONIC WEALTH", bold=True, size=16, color=NAVY)
cell(ws, "B3", "Portfolio Snapshot — NDPMS Review", bold=True, size=12, color=INDIGO)
cell(ws, "B5", "Client", bold=True, color=SLATE)
cell(ws, "C5", CLIENT, bold=True, size=11)
cell(ws, "B6", "As of", bold=True, color=SLATE)
cell(ws, "C6", AS_OF)
cell(ws, "B7", "Account type", bold=True, color=SLATE)
cell(ws, "C7", "NDPMS (Non-Discretionary PMS)")

# KPI band
cell(ws, "B9", "PORTFOLIO STRUCTURE", bold=True, color="FFFFFF", fill=NAVY, border=True)
for r in ("C9","D9","E9","F9"):
    cell(ws, r, "", fill=NAVY, border=True)

kpis = [
    ("Total holdings", "=COUNTA('Equity Holdings'!B7:B16)+COUNTA('Mutual Funds'!B7:B10)"),
    ("Stocks",         "=COUNTA('Equity Holdings'!B7:B16)"),
    ("Mutual funds",   "=COUNTA('Mutual Funds'!B7:B10)"),
    ("Weighting",      "Equal weight (all holdings)"),
]
r = 10
for lab, val in kpis:
    cell(ws, f"B{r}", lab, bold=True, color=SLATE, border=True, fill=NEAR)
    c = cell(ws, f"C{r}", val, border=True)
    r += 1

# Sleeve split (formula-driven)
cell(ws, "B15", "SLEEVE SPLIT", bold=True, color="FFFFFF", fill=NAVY, border=True)
for r in ("C15","D15"):
    cell(ws, r, "", fill=NAVY, border=True)
header_band(ws, 16, (2,), ["Sleeve"])
cell(ws, "C16", "Weight", bold=True, color="FFFFFF", fill=NAVY, align="center", border=True)
cell(ws, "D16", "Holdings", bold=True, color="FFFFFF", fill=NAVY, align="center", border=True)

cell(ws, "B17", "Equity (direct stocks)", border=True)
cell(ws, "C17", "=SUM('Equity Holdings'!D7:D16)", border=True, num="0.0%", align="center")
cell(ws, "D17", "=COUNTA('Equity Holdings'!B7:B16)", border=True, align="center")
cell(ws, "B18", "Mutual funds", border=True)
cell(ws, "C18", "=SUM('Mutual Funds'!E7:E10)", border=True, num="0.0%", align="center")
cell(ws, "D18", "=COUNTA('Mutual Funds'!B7:B10)", border=True, align="center")
cell(ws, "B19", "Total", bold=True, fill=TINT_A, border=True)
cell(ws, "C19", "=C17+C18", bold=True, fill=TINT_A, border=True, num="0.0%", align="center")
cell(ws, "D19", "=D17+D18", bold=True, fill=TINT_A, border=True, align="center")

# Watch-outs
cell(ws, "B21", "WATCH-OUTS (structure only)", bold=True, color="FFFFFF", fill=NAVY, border=True)
for r in ("C21","D21","E21","F21"):
    cell(ws, r, "", fill=NAVY, border=True)
notes = [
    "Two Small Cap funds held (LIC + Nippon) — category concentration in small-cap.",
    "IT sector doubled up (Infosys + KPIT Technologies) — see Sector Exposure sheet.",
    "LIC Small Cap is a Regular plan — cost-drag / switch-to-Direct candidate (confirm via fund framework).",
    "Scores & recommendations NOT filled — require Scorecard-750 + QFRA run (see Assumptions).",
]
r = 22
for n in notes:
    cell(ws, f"B{r}", "•", bold=True, color=AMBER)
    cell(ws, f"C{r}", n, wrap=True, size=9)
    ws.merge_cells(f"C{r}:F{r}")
    ws.row_dimensions[r].height = 26
    r += 1

# ============================================================ SHEET 2: EQUITY
we = wb.create_sheet("Equity Holdings")
we.sheet_view.showGridLines = False
widths = {"A":2, "B":16, "C":30, "D":11, "E":26, "F":14, "G":16, "H":22}
for col, w in widths.items():
    we.column_dimensions[col].width = w

cell(we, "B2", "EQUITY BOOK — Direct Stocks", bold=True, size=13, color=NAVY)
cell(we, "B3", f"{CLIENT}  ·  Equal weight  ·  as of {AS_OF}", color=SLATE, size=9, italic=True)

hdr = ["Ticker", "Company", "Weight", "Sector", "Ionic Score", "Recommendation", "Trim-to / Notes"]
header_band(we, 6, (2,), hdr)
we.row_dimensions[6].height = 30

row = 7
for tkr, name, sector in STOCKS:
    cell(we, f"B{row}", tkr, bold=True, border=True)
    cell(we, f"C{row}", name, border=True, size=9)
    # equal weight = 1/N_HOLD, stored as fraction
    cell(we, f"D{row}", f"=1/Summary!$C$10", border=True, num="0.0%", align="center")
    cell(we, f"E{row}", sector, border=True, size=9)
    cell(we, f"F{row}", "", border=True, align="center", fill=YELLOWFILL)   # pending
    cell(we, f"G{row}", "", border=True, align="center", fill=YELLOWFILL)   # pending
    cell(we, f"H{row}", "", border=True, size=9, fill=YELLOWFILL)          # pending
    row += 1

# total row
cell(we, f"B{row}", "TOTAL", bold=True, fill=TINT_A, border=True)
cell(we, f"C{row}", f"=COUNTA(B7:B{row-1})&\" stocks\"", fill=TINT_A, border=True, size=9)
cell(we, f"D{row}", f"=SUM(D7:D{row-1})", bold=True, fill=TINT_A, border=True, num="0.0%", align="center")
for col in ("E","F","G","H"):
    cell(we, f"{col}{row}", "", fill=TINT_A, border=True)
cell(we, f"F{row+2}", "Yellow cells: filled by the Scorecard-750 run.", italic=True, size=8, color=SLATE)

# ============================================================ SHEET 3: FUNDS
wf = wb.create_sheet("Mutual Funds")
wf.sheet_view.showGridLines = False
widths = {"A":2, "B":40, "C":18, "D":20, "E":11, "F":12, "G":16, "H":22}
for col, w in widths.items():
    wf.column_dimensions[col].width = w

cell(wf, "B2", "FUND BOOK — Mutual Funds", bold=True, size=13, color=NAVY)
cell(wf, "B3", f"{CLIENT}  ·  Equal weight  ·  as of {AS_OF}", color=SLATE, size=9, italic=True)

hdr = ["Scheme", "AMC", "Category", "Weight", "Plan", "Fund Score /100", "Verdict / Notes"]
header_band(wf, 6, (2,), hdr)
wf.row_dimensions[6].height = 30

row = 7
for scheme, amc, cat, plan in FUNDS:
    cell(wf, f"B{row}", scheme, bold=True, border=True, size=9, wrap=True)
    cell(wf, f"C{row}", amc, border=True, size=9)
    cell(wf, f"D{row}", cat, border=True, size=9)
    cell(wf, f"E{row}", f"=1/Summary!$C$10", border=True, num="0.0%", align="center")
    cell(wf, f"F{row}", plan, border=True, align="center", size=9)
    cell(wf, f"G{row}", "", border=True, align="center", fill=YELLOWFILL)   # pending
    cell(wf, f"H{row}", "", border=True, size=9, fill=YELLOWFILL)          # pending
    wf.row_dimensions[row].height = 24
    row += 1

cell(wf, f"B{row}", "TOTAL", bold=True, fill=TINT_A, border=True)
cell(wf, f"C{row}", f"=COUNTA(B7:B{row-1})&\" funds\"", fill=TINT_A, border=True, size=9)
cell(wf, f"D{row}", "", fill=TINT_A, border=True)
cell(wf, f"E{row}", f"=SUM(E7:E{row-1})", bold=True, fill=TINT_A, border=True, num="0.0%", align="center")
for col in ("F","G","H"):
    cell(wf, f"{col}{row}", "", fill=TINT_A, border=True)
cell(wf, f"B{row+2}", "Verdict rule: a fund Sell ships only when BOTH fund frameworks agree (else Hold).",
     italic=True, size=8, color=SLATE)
cell(wf, f"B{row+3}", "Yellow cells: filled by the fund-quality run.", italic=True, size=8, color=SLATE)

# ============================================================ SHEET 4: SECTOR
wsx = wb.create_sheet("Sector Exposure")
wsx.sheet_view.showGridLines = False
for col, w in {"A":2, "B":30, "C":14, "D":12}.items():
    wsx.column_dimensions[col].width = w
cell(wsx, "B2", "SECTOR EXPOSURE — Equity Sleeve", bold=True, size=13, color=NAVY)
cell(wsx, "B3", "Direct stocks only, by weight", color=SLATE, size=9, italic=True)

header_band(wsx, 6, (2,), ["Sector"])
cell(wsx, "C6", "Weight", bold=True, color="FFFFFF", fill=NAVY, align="center", border=True)
cell(wsx, "D6", "Count", bold=True, color="FFFFFF", fill=NAVY, align="center", border=True)

# unique sectors in first-seen order
seen = []
for _, _, s in STOCKS:
    if s not in seen:
        seen.append(s)
row = 7
for s in seen:
    cell(wsx, f"B{row}", s, border=True, size=9)
    cell(wsx, f"C{row}",
         f"=SUMIF('Equity Holdings'!$E$7:$E$16,B{row},'Equity Holdings'!$D$7:$D$16)",
         border=True, num="0.0%", align="center")
    cell(wsx, f"D{row}",
         f"=COUNTIF('Equity Holdings'!$E$7:$E$16,B{row})",
         border=True, align="center")
    row += 1
cell(wsx, f"B{row}", "Equity total", bold=True, fill=TINT_A, border=True)
cell(wsx, f"C{row}", f"=SUM(C7:C{row-1})", bold=True, fill=TINT_A, border=True, num="0.0%", align="center")
cell(wsx, f"D{row}", f"=SUM(D7:D{row-1})", bold=True, fill=TINT_A, border=True, align="center")

# ============================================================ SHEET 5: NOTES
wn = wb.create_sheet("Assumptions")
wn.sheet_view.showGridLines = False
wn.column_dimensions["A"].width = 2
wn.column_dimensions["B"].width = 100
cell(wn, "B2", "ASSUMPTIONS & DATA STATUS", bold=True, size=13, color=NAVY)
lines = [
    ("h", "Portfolio composition"),
    ("t", f"Client: {CLIENT}.  As-of date: {AS_OF}.  Account: NDPMS."),
    ("t", "You asked for '15 stocks / 12 stocks' but named 10 distinct companies (after merging"),
    ("t", "     RELIANCE and 'RELIANCE INDUSTRIES' into one holding — they are the same company)."),
    ("t", "     This file is built with those 10 real stocks. To reach 15, add the missing names and"),
    ("t", "     the equal weights will re-balance automatically (weight = 1 / total-holdings count)."),
    ("t", "4 mutual funds included exactly as named."),
    ("h", "Weighting"),
    ("t", "'All equal weight' read as equal weight across ALL holdings: each = 1 / total holdings."),
    ("t", "     With 10 stocks + 4 funds = 14 holdings, each holding = 7.14%; equity sleeve 71.4%,"),
    ("t", "     fund sleeve 28.6%. Change the split by editing weights or adding/removing rows."),
    ("t", "No total portfolio value (rupee amount) was provided, so the file is in % weights only."),
    ("t", "     Give me a total corpus and I will add rupee value columns per holding."),
    ("h", "What is NOT in this file (and why)"),
    ("t", "Ionic Scores, Buy/Sell/Hold/Trim recommendations, and fund quality verdicts are BLANK."),
    ("t", "     These are outputs of the Stock Scorecard 750 (quant + analyst research) and the fund"),
    ("t", "     quality frameworks. That pipeline needs market data + a research pass that has not been"),
    ("t", "     run here. Per firm rule, we never fabricate scores or calls — the yellow cells are"),
    ("t", "     placeholders for that run. Once scored, they drop straight into the yellow columns."),
    ("h", "Sector classifications"),
    ("t", "Sectors are standard market classifications for each company (e.g. Glenmark = Pharma,"),
    ("t", "     Anant Raj = Realty, Vodafone Idea = Telecom). Verify against your data feed if needed."),
    ("h", "Structural flags surfaced"),
    ("t", "Two Small Cap funds (LIC + Nippon) = small-cap category concentration to review."),
    ("t", "IT appears twice (Infosys + KPIT) = sector doubling in the equity sleeve."),
    ("t", "LIC Small Cap held in Regular plan = higher cost; Direct-plan switch is a common review item."),
]
row = 4
for kind, txt in lines:
    if kind == "h":
        cell(wn, f"B{row}", txt, bold=True, color=INDIGO, size=11)
    else:
        cell(wn, f"B{row}", txt, size=10, wrap=True)
    row += 1

# Force every spreadsheet app (Excel / Google Sheets / LibreOffice) to recalc
# all formulas on open. openpyxl writes no cached values, so this guarantees the
# reader sees computed numbers even though this sandbox cannot run LibreOffice.
wb.calculation.fullCalcOnLoad = True

out = "Shreyas_Gupta_Portfolio.xlsx"
wb.save(out)
print("saved", out)
